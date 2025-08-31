from __future__ import annotations
import csv
import json
from pathlib import Path
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from io import StringIO, BytesIO
import base64

from agents import RunContextWrapper, function_tool
from ..context import RunDeps

# For Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# For PDF support
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


@function_tool
def export_transactions(
    ctx: RunContextWrapper[RunDeps],
    format: str = "csv",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    output_file: Optional[str] = None
) -> str:
    """Export transactions to various formats for analysis or tax preparation.
    
    Args:
        format: Export format - "csv", "json", "excel", or "pdf"
        start_date: Optional start date (YYYY-MM-DD)
        end_date: Optional end date (YYYY-MM-DD)
        category: Optional category filter
        output_file: Optional output filename (auto-generated if not provided)
    """
    deps = ctx.context
    cur = deps.db.conn.cursor()
    
    # Build query
    query = "SELECT * FROM transactions WHERE 1=1"
    params = []
    
    if start_date:
        query += " AND date >= ?"
        params.append(start_date)
    
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
    
    if category:
        query += " AND category = ?"
        params.append(category.lower())
    
    query += " ORDER BY date DESC"
    
    cur.execute(query, params)
    transactions = cur.fetchall()
    
    if not transactions:
        return "No transactions found matching the criteria."
    
    # Convert to list of dicts
    tx_list = []
    for tx in transactions:
        tx_list.append({
            'id': tx['id'],
            'date': tx['date'],
            'description': tx['description'],
            'amount': tx['amount'],
            'currency': tx['currency'],
            'category': tx['category'] or '',
            'source_file': tx['source_file'] or ''
        })
    
    # Generate filename
    if not output_file:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"transactions_export_{timestamp}.{format}"
    
    export_path = deps.config.documents_dir / output_file
    
    # Export based on format
    if format == "csv":
        result = _export_csv(tx_list, export_path)
    elif format == "json":
        result = _export_json(tx_list, export_path)
    elif format == "excel":
        if not EXCEL_AVAILABLE:
            return "Excel export requires 'openpyxl'. Install with: pip install openpyxl"
        result = _export_excel(tx_list, export_path, deps)
    elif format == "pdf":
        if not PDF_AVAILABLE:
            return "PDF export requires 'reportlab'. Install with: pip install reportlab"
        result = _export_pdf(tx_list, export_path, deps)
    else:
        return f"Unsupported format: {format}. Use csv, json, excel, or pdf."
    
    return f"✅ Exported {len(tx_list)} transactions to {export_path}\n{result}"


def _export_csv(transactions: List[Dict], filepath: Path) -> str:
    """Export to CSV format."""
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        if transactions:
            writer = csv.DictWriter(f, fieldnames=transactions[0].keys())
            writer.writeheader()
            writer.writerows(transactions)
    
    total = sum(tx['amount'] for tx in transactions)
    return f"Format: CSV | Total amount: €{abs(total):.2f}"


def _export_json(transactions: List[Dict], filepath: Path) -> str:
    """Export to JSON format."""
    # Add metadata
    export_data = {
        'export_date': datetime.now().isoformat(),
        'total_transactions': len(transactions),
        'total_spent': sum(tx['amount'] for tx in transactions if tx['amount'] < 0),
        'total_received': sum(tx['amount'] for tx in transactions if tx['amount'] > 0),
        'transactions': transactions
    }
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
    
    return f"Format: JSON with metadata | Records: {len(transactions)}"


def _export_excel(transactions: List[Dict], filepath: Path, deps: RunDeps) -> str:
    """Export to Excel format with formatting."""
    wb = openpyxl.Workbook()
    
    # Transactions sheet
    ws = wb.active
    ws.title = "Transactions"
    
    # Headers with formatting
    headers = ['ID', 'Date', 'Description', 'Amount', 'Currency', 'Category', 'Source']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, tx in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=tx['id'])
        ws.cell(row=row, column=2, value=tx['date'])
        ws.cell(row=row, column=3, value=tx['description'])
        
        # Format amount with color
        amount_cell = ws.cell(row=row, column=4, value=tx['amount'])
        if tx['amount'] < 0:
            amount_cell.font = Font(color="FF0000")  # Red for expenses
        else:
            amount_cell.font = Font(color="008000")  # Green for income
        
        ws.cell(row=row, column=5, value=tx['currency'])
        ws.cell(row=row, column=6, value=tx['category'])
        ws.cell(row=row, column=7, value=tx['source_file'])
    
    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Calculate summaries
    total_spent = sum(tx['amount'] for tx in transactions if tx['amount'] < 0)
    total_received = sum(tx['amount'] for tx in transactions if tx['amount'] > 0)
    categories = {}
    for tx in transactions:
        if tx['category'] and tx['amount'] < 0:
            categories[tx['category']] = categories.get(tx['category'], 0) + abs(tx['amount'])
    
    # Write summary
    summary_data = [
        ['Summary Report', ''],
        ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['', ''],
        ['Total Transactions', len(transactions)],
        ['Total Spent', f'€{abs(total_spent):.2f}'],
        ['Total Received', f'€{total_received:.2f}'],
        ['Net', f'€{total_received + total_spent:.2f}'],
        ['', ''],
        ['Top Categories', 'Amount']
    ]
    
    for category, amount in sorted(categories.items(), key=lambda x: x[1], reverse=True)[:10]:
        summary_data.append([category.title(), f'€{amount:.2f}'])
    
    for row, data in enumerate(summary_data, 1):
        for col, value in enumerate(data, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if row == 1:
                cell.font = Font(bold=True, size=14)
            elif row == 9:
                cell.font = Font(bold=True)
    
    # Budget comparison sheet if budgets exist
    cur = deps.db.conn.cursor()
    cur.execute("SELECT * FROM budgets")
    budgets = cur.fetchall()
    
    if budgets:
        ws_budget = wb.create_sheet("Budget Analysis")
        budget_headers = ['Category', 'Budget', 'Spent', 'Remaining', 'Status']
        
        for col, header in enumerate(budget_headers, 1):
            cell = ws_budget.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, budget in enumerate(budgets, 2):
            spent = categories.get(budget['category'], 0)
            remaining = budget['amount'] - spent
            status = "✅ Under" if remaining > 0 else "⚠️ Over"
            
            ws_budget.cell(row=row, column=1, value=budget['category'].title())
            ws_budget.cell(row=row, column=2, value=f"€{budget['amount']:.2f}")
            ws_budget.cell(row=row, column=3, value=f"€{spent:.2f}")
            ws_budget.cell(row=row, column=4, value=f"€{remaining:.2f}")
            ws_budget.cell(row=row, column=5, value=status)
    
    wb.save(filepath)
    return f"Format: Excel with {len(wb.worksheets)} sheets | Enhanced with formatting and summaries"


def _export_pdf(transactions: List[Dict], filepath: Path, deps: RunDeps) -> str:
    """Export to PDF format with formatting."""
    doc = SimpleDocTemplate(str(filepath), pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=30
    )
    elements.append(Paragraph("Financial Transaction Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Metadata
    metadata = [
        f"<b>Report Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"<b>Total Transactions:</b> {len(transactions)}",
        f"<b>Date Range:</b> {transactions[-1]['date']} to {transactions[0]['date']}" if transactions else ""
    ]
    
    for meta in metadata:
        elements.append(Paragraph(meta, styles['Normal']))
    
    elements.append(Spacer(1, 20))
    
    # Summary statistics
    total_spent = sum(tx['amount'] for tx in transactions if tx['amount'] < 0)
    total_received = sum(tx['amount'] for tx in transactions if tx['amount'] > 0)
    
    summary_data = [
        ['Summary', 'Amount'],
        ['Total Spent', f'€{abs(total_spent):.2f}'],
        ['Total Received', f'€{total_received:.2f}'],
        ['Net Amount', f'€{total_received + total_spent:.2f}']
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(PageBreak())
    
    # Transaction table (paginated)
    elements.append(Paragraph("Transaction Details", styles['Heading1']))
    elements.append(Spacer(1, 12))
    
    # Prepare transaction data for table
    tx_data = [['Date', 'Description', 'Amount', 'Category']]
    
    for tx in transactions[:100]:  # Limit to 100 for PDF size
        desc = tx['description'][:40] + '...' if len(tx['description']) > 40 else tx['description']
        tx_data.append([
            tx['date'],
            desc,
            f"€{tx['amount']:.2f}",
            tx['category'] or '-'
        ])
    
    # Create transaction table
    tx_table = Table(tx_data, colWidths=[1.5*inch, 3*inch, 1.5*inch, 1.5*inch])
    tx_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),  # Right-align amounts
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    # Color negative amounts
    for i, tx in enumerate(transactions[:100], 1):
        if tx['amount'] < 0:
            tx_table.setStyle(TableStyle([
                ('TEXTCOLOR', (2, i), (2, i), colors.red)
            ]))
    
    elements.append(tx_table)
    
    if len(transactions) > 100:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"<i>Note: Showing first 100 of {len(transactions)} transactions</i>", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    
    return f"Format: PDF Report | Pages: ~{len(transactions)//30 + 2} | Professional layout"


@function_tool
def generate_tax_report(
    ctx: RunContextWrapper[RunDeps],
    year: Optional[int] = None,
    format: str = "pdf"
) -> str:
    """Generate a comprehensive tax report with categorized expenses.
    
    Args:
        year: Tax year (defaults to previous year)
        format: Output format - "pdf", "excel", or "csv"
    """
    deps = ctx.context
    cur = deps.db.conn.cursor()
    
    # Determine tax year
    if not year:
        year = datetime.now().year - 1 if datetime.now().month < 4 else datetime.now().year
    
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    # Get all transactions for the year
    cur.execute(
        """SELECT * FROM transactions 
           WHERE date >= ? AND date <= ?
           ORDER BY date, category""",
        (start_date, end_date)
    )
    transactions = cur.fetchall()
    
    if not transactions:
        return f"No transactions found for tax year {year}"
    
    # Categorize for tax purposes
    tax_categories = {
        'income': [],
        'business_expenses': [],
        'medical': [],
        'charity': [],
        'education': [],
        'home_office': [],
        'travel': [],
        'other_deductions': []
    }
    
    # Common tax-deductible categories mapping
    category_mapping = {
        'salary': 'income',
        'freelance': 'income',
        'business': 'business_expenses',
        'office': 'home_office',
        'medical': 'medical',
        'health': 'medical',
        'donation': 'charity',
        'education': 'education',
        'training': 'education',
        'travel': 'travel',
        'transport': 'travel'
    }
    
    # Categorize transactions
    for tx in transactions:
        tx_dict = dict(tx)
        
        if tx['amount'] > 0:
            tax_categories['income'].append(tx_dict)
        elif tx['category']:
            assigned = False
            for keyword, tax_cat in category_mapping.items():
                if keyword in tx['category'].lower():
                    tax_categories[tax_cat].append(tx_dict)
                    assigned = True
                    break
            
            if not assigned and tx['amount'] < 0:
                tax_categories['other_deductions'].append(tx_dict)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"tax_report_{year}_{timestamp}.{format}"
    filepath = deps.config.documents_dir / filename
    
    if format == "pdf":
        if not PDF_AVAILABLE:
            return "PDF generation requires 'reportlab'. Install with: pip install reportlab"
        result = _generate_tax_pdf(tax_categories, year, filepath)
    elif format == "excel":
        if not EXCEL_AVAILABLE:
            return "Excel generation requires 'openpyxl'. Install with: pip install openpyxl"
        result = _generate_tax_excel(tax_categories, year, filepath)
    elif format == "csv":
        result = _generate_tax_csv(tax_categories, year, filepath)
    else:
        return f"Unsupported format: {format}"
    
    # Calculate totals
    total_income = sum(tx['amount'] for tx in tax_categories['income'])
    total_deductions = sum(
        abs(tx['amount']) 
        for cat in tax_categories.values() 
        if cat != tax_categories['income']
        for tx in cat
    )
    
    return f"""✅ Tax Report Generated for {year}
    
📊 Summary:
• Total Income: €{total_income:.2f}
• Total Deductions: €{total_deductions:.2f}
• Net Taxable: €{total_income - total_deductions:.2f}

📁 File: {filepath}
{result}"""


def _generate_tax_pdf(tax_categories: Dict, year: int, filepath: Path) -> str:
    """Generate PDF tax report."""
    doc = SimpleDocTemplate(str(filepath), pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'TaxTitle',
        parent=styles['Title'],
        fontSize=28,
        textColor=colors.HexColor('#003366'),
        alignment=1  # Center
    )
    elements.append(Paragraph(f"Tax Report {year}", title_style))
    elements.append(Spacer(1, 30))
    
    # Summary table
    total_income = sum(tx['amount'] for tx in tax_categories['income'])
    total_deductions = sum(
        abs(tx['amount']) 
        for cat_name, cat in tax_categories.items() 
        if cat_name != 'income'
        for tx in cat
    )
    
    summary_data = [
        ['Tax Summary', 'Amount (€)'],
        ['Gross Income', f'{total_income:.2f}'],
        ['Total Deductions', f'{total_deductions:.2f}'],
        ['Net Taxable Income', f'{total_income - total_deductions:.2f}']
    ]
    
    summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(PageBreak())
    
    # Detailed sections
    for category_name, transactions in tax_categories.items():
        if not transactions:
            continue
        
        # Section header
        elements.append(Paragraph(category_name.replace('_', ' ').title(), styles['Heading1']))
        elements.append(Spacer(1, 12))
        
        # Transaction details
        cat_data = [['Date', 'Description', 'Amount']]
        cat_total = 0
        
        for tx in transactions[:50]:  # Limit per category
            amount = abs(tx['amount'])
            cat_total += amount
            cat_data.append([
                tx['date'],
                tx['description'][:50],
                f'€{amount:.2f}'
            ])
        
        # Add total row
        cat_data.append(['', 'TOTAL', f'€{cat_total:.2f}'])
        
        cat_table = Table(cat_data, colWidths=[1.5*inch, 3.5*inch, 1.5*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(cat_table)
        
        if len(transactions) > 50:
            elements.append(Paragraph(f"<i>({len(transactions)-50} more transactions not shown)</i>", styles['Normal']))
        
        elements.append(Spacer(1, 20))
    
    doc.build(elements)
    return "Professional PDF tax report with categorized deductions"


def _generate_tax_excel(tax_categories: Dict, year: int, filepath: Path) -> str:
    """Generate Excel tax report."""
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Tax Summary"
    
    # Calculate totals
    total_income = sum(tx['amount'] for tx in tax_categories['income'])
    total_by_category = {}
    
    for cat_name, transactions in tax_categories.items():
        if cat_name != 'income' and transactions:
            total_by_category[cat_name] = sum(abs(tx['amount']) for tx in transactions)
    
    # Write summary
    ws_summary['A1'] = f'Tax Report {year}'
    ws_summary['A1'].font = Font(size=20, bold=True)
    
    row = 3
    ws_summary[f'A{row}'] = 'Category'
    ws_summary[f'B{row}'] = 'Amount (€)'
    ws_summary[f'A{row}:B{row}'][0].font = Font(bold=True)
    
    row += 1
    ws_summary[f'A{row}'] = 'Gross Income'
    ws_summary[f'B{row}'] = total_income
    
    for cat_name, total in total_by_category.items():
        row += 1
        ws_summary[f'A{row}'] = cat_name.replace('_', ' ').title()
        ws_summary[f'B{row}'] = -total
    
    row += 2
    ws_summary[f'A{row}'] = 'Net Taxable Income'
    ws_summary[f'B{row}'] = total_income - sum(total_by_category.values())
    ws_summary[f'A{row}:B{row}'][0].font = Font(bold=True)
    
    # Create sheet for each category with transactions
    for cat_name, transactions in tax_categories.items():
        if not transactions:
            continue
        
        ws = wb.create_sheet(cat_name.replace('_', ' ').title())
        
        # Headers
        headers = ['Date', 'Description', 'Amount', 'Category']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        # Data
        for row, tx in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=tx['date'])
            ws.cell(row=row, column=2, value=tx['description'])
            ws.cell(row=row, column=3, value=abs(tx['amount']))
            ws.cell(row=row, column=4, value=tx.get('category', ''))
    
    wb.save(filepath)
    return f"Excel workbook with {len(wb.worksheets)} categorized sheets"


def _generate_tax_csv(tax_categories: Dict, year: int, filepath: Path) -> str:
    """Generate CSV tax report."""
    all_transactions = []
    
    for cat_name, transactions in tax_categories.items():
        for tx in transactions:
            all_transactions.append({
                'tax_category': cat_name,
                'date': tx['date'],
                'description': tx['description'],
                'amount': tx['amount'],
                'original_category': tx.get('category', '')
            })
    
    # Sort by date
    all_transactions.sort(key=lambda x: x['date'])
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        if all_transactions:
            writer = csv.DictWriter(f, fieldnames=all_transactions[0].keys())
            writer.writeheader()
            writer.writerows(all_transactions)
    
    return f"CSV with {len(all_transactions)} categorized transactions"


@function_tool  
def export_budget_report(
    ctx: RunContextWrapper[RunDeps],
    format: str = "pdf",
    period_days: int = 30
) -> str:
    """Export a comprehensive budget performance report.
    
    Args:
        format: Output format - "pdf", "excel", or "json"
        period_days: Number of days to analyze
    """
    deps = ctx.context
    cur = deps.db.conn.cursor()
    
    # Get all budgets
    cur.execute("SELECT * FROM budgets ORDER BY category")
    budgets = cur.fetchall()
    
    if not budgets:
        return "No budgets found. Set budgets first using set_budget."
    
    # Calculate spending for each budget
    end_date = datetime.now()
    start_date = end_date - timedelta(days=period_days)
    
    budget_data = []
    for budget in budgets:
        # Get spending
        cur.execute(
            """SELECT COALESCE(SUM(ABS(amount)), 0) as spent
               FROM transactions 
               WHERE category = ? 
               AND amount < 0
               AND date >= ?
               AND date <= ?""",
            (budget['category'], start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
        )
        spent = cur.fetchone()['spent']
        
        # Adjust budget based on period
        if budget['period'] == 'weekly':
            adjusted_budget = budget['amount'] * (period_days / 7)
        elif budget['period'] == 'yearly': 
            adjusted_budget = budget['amount'] * (period_days / 365)
        else:  # monthly
            adjusted_budget = budget['amount'] * (period_days / 30)
        
        budget_data.append({
            'category': budget['category'],
            'budget_amount': adjusted_budget,
            'spent': spent,
            'remaining': adjusted_budget - spent,
            'percentage': (spent / adjusted_budget * 100) if adjusted_budget > 0 else 0,
            'period': budget['period'],
            'status': 'over' if spent > adjusted_budget else 'under'
        })
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"budget_report_{timestamp}.{format}"
    filepath = deps.config.documents_dir / filename
    
    # Export based on format
    if format == "json":
        export_data = {
            'report_date': datetime.now().isoformat(),
            'period_days': period_days,
            'date_range': {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d')
            },
            'budgets': budget_data,
            'summary': {
                'total_budget': sum(b['budget_amount'] for b in budget_data),
                'total_spent': sum(b['spent'] for b in budget_data),
                'total_remaining': sum(b['remaining'] for b in budget_data),
                'categories_over_budget': sum(1 for b in budget_data if b['status'] == 'over')
            }
        }
        
        with open(filepath, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        result = "JSON budget report with full metrics"
    
    elif format == "pdf" and PDF_AVAILABLE:
        doc = SimpleDocTemplate(str(filepath), pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        elements.append(Paragraph("Budget Performance Report", styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Budget table
        table_data = [['Category', 'Budget', 'Spent', 'Remaining', 'Status']]
        
        for b in budget_data:
            status_icon = "🔴" if b['status'] == 'over' else "🟢"
            table_data.append([
                b['category'].title(),
                f"€{b['budget_amount']:.2f}",
                f"€{b['spent']:.2f}",
                f"€{b['remaining']:.2f}",
                f"{b['percentage']:.0f}% {status_icon}"
            ])
        
        budget_table = Table(table_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        budget_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(budget_table)
        doc.build(elements)
        result = "PDF budget performance report"
    
    elif format == "excel" and EXCEL_AVAILABLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget Report"
        
        # Headers
        headers = ['Category', 'Budget Period', 'Budget Amount', 'Spent', 'Remaining', 'Usage %', 'Status']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        # Data
        for row, b in enumerate(budget_data, 2):
            ws.cell(row=row, column=1, value=b['category'].title())
            ws.cell(row=row, column=2, value=b['period'])
            ws.cell(row=row, column=3, value=b['budget_amount'])
            ws.cell(row=row, column=4, value=b['spent'])
            ws.cell(row=row, column=5, value=b['remaining'])
            ws.cell(row=row, column=6, value=f"{b['percentage']:.1f}%")
            ws.cell(row=row, column=7, value=b['status'].upper())
            
            # Color code status
            if b['status'] == 'over':
                ws.cell(row=row, column=7).font = Font(color="FF0000")
        
        wb.save(filepath)
        result = "Excel budget report with status indicators"
    else:
        return f"Format {format} not available or not installed"
    
    return f"✅ Budget report exported to {filepath}\n{result}"


@function_tool
def export_recurring_payments(
    ctx: RunContextWrapper[RunDeps],
    format: str = "pdf",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    min_confidence: float = 0.6,
    exclude_credit_repayment: bool = True,
    bills_only: bool = False
) -> str:
    """Export only recurring payments/subscriptions to a clean report.
    
    Args:
        format: Export format - "pdf", "excel", "csv", or "json"
        start_date: Optional start date filter (YYYY-MM-DD)
        end_date: Optional end date filter (YYYY-MM-DD) 
        min_confidence: Minimum confidence level for recurring detection (0.0-1.0)
    """
    deps = ctx.context
    cur = deps.db.conn.cursor()
    
    # Check if recurring patterns exist in database first
    cur.execute("SELECT COUNT(*) as count FROM recurring_transactions WHERE confidence >= ?", (min_confidence,))
    existing_patterns = cur.fetchone()['count']
    
    if existing_patterns == 0:
        return f"""No recurring transactions detected with confidence >= {min_confidence*100:.0f}%.

To export recurring payments, you need to detect them first:
1. Run 'detect_recurring' to analyze your transactions for recurring patterns
2. Then use this export function to generate reports

Example: "Please detect recurring transactions first, then export them to PDF"
"""
    
    # Get recurring transactions from database
    cur.execute(
        """SELECT * FROM recurring_transactions 
           WHERE confidence >= ?
           ORDER BY amount DESC, confidence DESC""",
        (min_confidence,)
    )
    recurring_patterns = cur.fetchall()

    # Optional filter: exclude credit card repayment pattern and POS-like patterns for bills_only
    def is_credit_repayment(p) -> bool:
        d = (p['description_pattern'] or '').lower()
        return ('creditcard' in d or 'credit card' in d) and ('incasso' in d or 'afloss' in d or 'accountnr' in d)

    if exclude_credit_repayment:
        recurring_patterns = [p for p in recurring_patterns if not is_credit_repayment(p)]

    if bills_only:
        def is_pos_like(p) -> bool:
            d = (p['description_pattern'] or '').lower()
            return any(x in d for x in [
                'betaalautomaat', 'geldmaat', 'albert heijn', ' ah ', 'jumbo', 'primera', 'coffeeshop', 'restaurant', 'mc donald', 'kfc'
            ])
        # Keep likely monthly bills; drop obvious POS and fees/cash withdrawals
        recurring_patterns = [
            p for p in recurring_patterns
            if not is_pos_like(p) and (p['category'] or '').lower() not in ('fees', 'cash_withdrawal')
        ]
    
    if not recurring_patterns:
        return f"No recurring transactions found with confidence >= {min_confidence*100:.0f}%"
    
    # Get actual transactions matching these patterns within date range
    query = """
        SELECT t.* FROM transactions t
        WHERE t.amount < 0
    """
    params = []
    
    if start_date:
        query += " AND t.date >= ?"
        params.append(start_date)
    
    if end_date:
        query += " AND t.date <= ?"
        params.append(end_date)
    
    query += " ORDER BY t.date DESC"
    
    cur.execute(query, params)
    all_transactions = cur.fetchall()
    
    # Filter to only transactions that match recurring patterns
    recurring_transactions = []
    
    # Whitelist keywords for variable bills
    bill_keywords = [
        # telecom
        'vodafone', 'libertel', 'odido', 'kpn', 't-mobile', 'tmobile', 'ziggo',
        # energy/utilities
        'essent', 'energie', 'vandebron', 'nuon', 'eneco', 'waterbedrijf', 'stroom', 'gas',
        # insurance
        'vgz', 'verzekering', 'verzekeraar', 'nn schadeverzekering', 'anwb verzekeren', 'cz', 'fbto',
        # housing/rent
        'brabantwonen', 'huur', 'hypotheek', 'woning', 'woonverzekering',
        # government/taxes
        'gemeente', 'belasting', 'heffing', 'waterschap'
    ]

    def is_variable_bill(desc: str) -> bool:
        dl = desc.lower()
        return any(k in dl for k in bill_keywords)

    for tx in all_transactions:
        for pattern in recurring_patterns:
            # Simple pattern matching - check if transaction description contains key parts
            pattern_words = pattern['description_pattern'].lower().split()[:3]  # First 3 words
            tx_desc = tx['description'].lower()
            
            # Check if most pattern words are in transaction description
            matches = sum(1 for word in pattern_words if word in tx_desc)
            if matches >= len(pattern_words) * 0.7:  # 70% word match threshold
                # Amount tolerance: allow higher variance for variable bills
                tol = 0.2
                if bills_only and (is_variable_bill(tx_desc) or is_variable_bill(pattern['description_pattern'])):
                    tol = 0.4
                # Check if amount is similar
                amount_diff = abs(abs(tx['amount']) - pattern['amount']) / max(pattern['amount'], 1e-9)
                if amount_diff <= tol:
                    # bills_only extra filter against POS-like lines
                    if bills_only:
                        tcat = (tx['category'] or '').lower() if 'category' in tx.keys() else ''
                        if tcat in ('fees', 'cash_withdrawal'):
                            continue
                        tdl = tx_desc
                        if any(x in tdl for x in ['betaalautomaat', 'geldmaat', 'albert heijn', ' ah ', 'jumbo', 'coffeeshop']):
                            continue
                    tx_dict = dict(tx)
                    tx_dict['recurring_pattern'] = pattern['description_pattern']
                    tx_dict['confidence'] = pattern['confidence']
                    tx_dict['frequency'] = pattern['frequency']
                    recurring_transactions.append(tx_dict)
                    break
    
    if not recurring_transactions:
        return "No actual transactions found matching recurring patterns in the specified date range."
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    date_suffix = ""
    if start_date and end_date:
        date_suffix = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_suffix = f"_from_{start_date}"
    elif end_date:
        date_suffix = f"_until_{end_date}"
    
    filename = f"recurring_payments{date_suffix}_{timestamp}.{format}"
    filepath = deps.config.documents_dir / filename
    
    # Export based on format
    if format == "pdf":
        if not PDF_AVAILABLE:
            return "PDF export requires 'reportlab'. Install with: pip install reportlab"
        result = _export_recurring_pdf(recurring_transactions, recurring_patterns, filepath, start_date, end_date)
    elif format == "excel":
        if not EXCEL_AVAILABLE:
            return "Excel export requires 'openpyxl'. Install with: pip install openpyxl"
        result = _export_recurring_excel(recurring_transactions, recurring_patterns, filepath, start_date, end_date)
    elif format == "csv":
        result = _export_recurring_csv(recurring_transactions, filepath)
    elif format == "json":
        result = _export_recurring_json(recurring_transactions, recurring_patterns, filepath, start_date, end_date)
    else:
        return f"Unsupported format: {format}. Use pdf, excel, csv, or json."
    
    return f"✅ Exported {len(recurring_transactions)} recurring payments to {filepath}\n{result}"


def _export_recurring_pdf(transactions: List[Dict], patterns: List[Dict], filepath: Path, start_date: Optional[str] = None, end_date: Optional[str] = None) -> str:
    """Export recurring payments to PDF format."""
    doc = SimpleDocTemplate(str(filepath), pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'RecurringTitle',
        parent=styles['Title'],
        fontSize=24,
        textColor=colors.HexColor('#1f4788'),
        alignment=1
    )
    elements.append(Paragraph("Recurring Payments Report", title_style))
    elements.append(Spacer(1, 20))
    
    # Date range
    if start_date or end_date:
        date_range = f"Period: {start_date or 'All time'} to {end_date or 'Present'}"
        elements.append(Paragraph(date_range, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # Summary
    total_recurring = sum(abs(tx['amount']) for tx in transactions)
    unique_patterns = len(set(tx['recurring_pattern'] for tx in transactions))
    
    summary_data = [
        ['Summary', 'Value'],
        ['Total Recurring Payments', f'€{total_recurring:.2f}'],
        ['Number of Transactions', str(len(transactions))],
        ['Unique Recurring Patterns', str(unique_patterns)]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Group by recurring pattern
    by_pattern = {}
    for tx in transactions:
        pattern = tx['recurring_pattern']
        if pattern not in by_pattern:
            by_pattern[pattern] = []
        by_pattern[pattern].append(tx)
    
    # Pattern analysis table
    elements.append(Paragraph("Recurring Payment Analysis", styles['Heading1']))
    elements.append(Spacer(1, 12))
    
    pattern_data = [['Payment Type', 'Frequency', 'Avg Amount', 'Total', 'Count']]
    
    for pattern, pattern_txs in sorted(by_pattern.items(), key=lambda x: sum(abs(t['amount']) for t in x[1]), reverse=True):
        avg_amount = sum(abs(tx['amount']) for tx in pattern_txs) / len(pattern_txs)
        total_amount = sum(abs(tx['amount']) for tx in pattern_txs)
        frequency = pattern_txs[0]['frequency']
        
        pattern_data.append([
            pattern[:35],
            frequency.title(),
            f'€{avg_amount:.2f}',
            f'€{total_amount:.2f}',
            str(len(pattern_txs))
        ])
    
    pattern_table = Table(pattern_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 0.7*inch])
    pattern_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(pattern_table)
    
    doc.build(elements)
    return f"Professional PDF report with {len(by_pattern)} recurring payment types"


def _export_recurring_excel(transactions: List[Dict], patterns: List[Dict], filepath: Path, start_date: Optional[str] = None, end_date: Optional[str] = None) -> str:
    """Export recurring payments to Excel format."""
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Recurring Summary"
    
    # Header
    ws_summary['A1'] = 'Recurring Payments Report'
    ws_summary['A1'].font = Font(size=18, bold=True)
    
    if start_date or end_date:
        ws_summary['A2'] = f"Period: {start_date or 'All time'} to {end_date or 'Present'}"
        ws_summary['A2'].font = Font(size=12, italic=True)
    
    # Summary stats
    total_recurring = sum(abs(tx['amount']) for tx in transactions)
    unique_patterns = len(set(tx['recurring_pattern'] for tx in transactions))
    
    row = 4
    ws_summary[f'A{row}'] = 'Total Recurring Payments:'
    ws_summary[f'B{row}'] = f'€{total_recurring:.2f}'
    ws_summary[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws_summary[f'A{row}'] = 'Number of Transactions:'
    ws_summary[f'B{row}'] = len(transactions)
    
    row += 1
    ws_summary[f'A{row}'] = 'Unique Payment Types:'
    ws_summary[f'B{row}'] = unique_patterns
    
    # Group by pattern for analysis
    by_pattern = {}
    for tx in transactions:
        pattern = tx['recurring_pattern']
        if pattern not in by_pattern:
            by_pattern[pattern] = []
        by_pattern[pattern].append(tx)
    
    # Pattern analysis
    row += 3
    ws_summary[f'A{row}'] = 'Payment Type'
    ws_summary[f'B{row}'] = 'Frequency'
    ws_summary[f'C{row}'] = 'Avg Amount'
    ws_summary[f'D{row}'] = 'Total Amount'
    ws_summary[f'E{row}'] = 'Count'
    
    # Header formatting
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_summary[f'{col}{row}'].font = Font(bold=True)
        ws_summary[f'{col}{row}'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for pattern, pattern_txs in sorted(by_pattern.items(), key=lambda x: sum(abs(t['amount']) for t in x[1]), reverse=True):
        row += 1
        avg_amount = sum(abs(tx['amount']) for tx in pattern_txs) / len(pattern_txs)
        total_amount = sum(abs(tx['amount']) for tx in pattern_txs)
        
        ws_summary[f'A{row}'] = pattern
        ws_summary[f'B{row}'] = pattern_txs[0]['frequency'].title()
        ws_summary[f'C{row}'] = avg_amount
        ws_summary[f'D{row}'] = total_amount
        ws_summary[f'E{row}'] = len(pattern_txs)
    
    # All transactions sheet
    ws_detail = wb.create_sheet("Transaction Details")
    
    headers = ['Date', 'Description', 'Amount', 'Recurring Pattern', 'Frequency', 'Confidence']
    for col, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, tx in enumerate(transactions, 2):
        ws_detail.cell(row=row, column=1, value=tx['date'])
        ws_detail.cell(row=row, column=2, value=tx['description'])
        ws_detail.cell(row=row, column=3, value=abs(tx['amount']))
        ws_detail.cell(row=row, column=4, value=tx['recurring_pattern'])
        ws_detail.cell(row=row, column=5, value=tx['frequency'])
        ws_detail.cell(row=row, column=6, value=f"{tx['confidence']*100:.0f}%")
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_detail]:
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
    
    wb.save(filepath)
    return f"Excel workbook with summary and {len(transactions)} transaction details"


def _export_recurring_csv(transactions: List[Dict], filepath: Path) -> str:
    """Export recurring payments to CSV format."""
    csv_data = []
    for tx in transactions:
        csv_data.append({
            'date': tx['date'],
            'description': tx['description'],
            'amount': abs(tx['amount']),
            'currency': tx.get('currency', 'EUR'),
            'recurring_pattern': tx['recurring_pattern'],
            'frequency': tx['frequency'],
            'confidence': f"{tx['confidence']*100:.0f}%"
        })
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        if csv_data:
            writer = csv.DictWriter(f, fieldnames=csv_data[0].keys())
            writer.writeheader()
            writer.writerows(csv_data)
    
    total = sum(tx['amount'] for tx in csv_data)
    return f"CSV with {len(csv_data)} recurring transactions | Total: €{total:.2f}"


def _export_recurring_json(transactions: List[Dict], patterns: List[Dict], filepath: Path, start_date: Optional[str] = None, end_date: Optional[str] = None) -> str:
    """Export recurring payments to JSON format."""
    # Group by pattern
    by_pattern = {}
    for tx in transactions:
        pattern = tx['recurring_pattern']
        if pattern not in by_pattern:
            by_pattern[pattern] = []
        by_pattern[pattern].append(tx)
    
    # Build export data
    export_data = {
        'export_date': datetime.now().isoformat(),
        'date_range': {
            'start': start_date,
            'end': end_date
        },
        'summary': {
            'total_transactions': len(transactions),
            'total_amount': sum(abs(tx['amount']) for tx in transactions),
            'unique_patterns': len(by_pattern),
            'patterns': []
        },
        'transactions': transactions
    }
    
    # Add pattern summaries
    for pattern, pattern_txs in sorted(by_pattern.items(), key=lambda x: sum(abs(t['amount']) for t in x[1]), reverse=True):
        export_data['summary']['patterns'].append({
            'name': pattern,
            'frequency': pattern_txs[0]['frequency'],
            'transaction_count': len(pattern_txs),
            'total_amount': sum(abs(tx['amount']) for tx in pattern_txs),
            'average_amount': sum(abs(tx['amount']) for tx in pattern_txs) / len(pattern_txs),
            'confidence': pattern_txs[0]['confidence']
        })
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
    
    return f"JSON with metadata and {len(by_pattern)} recurring payment patterns"
